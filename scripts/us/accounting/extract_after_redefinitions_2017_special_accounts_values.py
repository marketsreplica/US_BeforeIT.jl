#!/usr/bin/env python3

"""Extract the four pinned BEA 2017 worksheets for the fixture canonicalizer."""

import hashlib
import json
from pathlib import Path
import platform
import sys

import openpyxl
from openpyxl import load_workbook


EXPECTED_OPENPYXL_VERSION = "3.1.5"
EXPECTED = {
    "IOUse_After_Redefinitions_PRO_Detail.xlsx": (
        2_113_284,
        "ee0f977ccc6b884d3e3b912596e39c1036f513880531dda33be947e68fb03fe4",
    ),
    "IOUse_After_Redefinitions_PRO_Summary.xlsx": (
        1_163_798,
        "9e3791d657909843ce202161bae00cf8a425d7e1bf866cc8a0462810f0ae00c7",
    ),
    "IOMake_After_Redefinitions_PRO_Detail.xlsx": (
        1_472_556,
        "96fb70a032e3ab81514231f49c2eae888b7ef8b741b00f352f2fc0fa8776db67",
    ),
    "IOMake_After_Redefinitions_PRO_Summary.xlsx": (
        598_989,
        "073b87c7e52e76fb78ad7ddafb0c2e60f9188fc5a4e56dc0094f4a7ae3f529c6",
    ),
}


def sha256(path: Path) -> str:
    return hashlib.sha256(path.read_bytes()).hexdigest()


def main() -> None:
    if len(sys.argv) != 3:
        raise SystemExit(
            "usage: extract_after_redefinitions_2017_special_accounts_values.py "
            "<source-directory> <output-json>"
        )
    if openpyxl.__version__ != EXPECTED_OPENPYXL_VERSION:
        raise RuntimeError(
            "openpyxl version changed: "
            f"{openpyxl.__version__} != {EXPECTED_OPENPYXL_VERSION}"
        )

    source_directory = Path(sys.argv[1])
    output_path = Path(sys.argv[2])
    workbooks = {}
    for member, (expected_bytes, expected_sha256) in EXPECTED.items():
        source_path = source_directory / member
        if source_path.stat().st_size != expected_bytes:
            raise RuntimeError(f"{member} byte count changed")
        if sha256(source_path) != expected_sha256:
            raise RuntimeError(f"{member} SHA-256 changed")

        workbook = load_workbook(source_path, read_only=True, data_only=True)
        worksheet = workbook["2017"]
        workbooks[member] = [
            [cell.value for cell in row]
            for row in worksheet.iter_rows(
                min_row=1,
                max_row=worksheet.max_row,
                min_col=1,
                max_col=worksheet.max_column,
            )
        ]
        workbook.close()

    output = {
        "generator_backend": "openpyxl",
        "python_version": platform.python_version(),
        "openpyxl_version": openpyxl.__version__,
        "workbooks": workbooks,
    }
    output_path.write_text(
        json.dumps(output, ensure_ascii=False, separators=(",", ":")),
        encoding="utf-8",
    )


if __name__ == "__main__":
    main()
