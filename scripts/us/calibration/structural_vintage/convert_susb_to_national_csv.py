#!/usr/bin/env python3
"""Convert a Census SUSB workbook to the national-total CSV schema the US BeforeIT
ingestion already checks (the 2022 us_state_6digitnaics txt layout):

    STATE,NAICS,ENTRSIZE,FIRM,ESTB,EMPL

Only national (State 00) "01: Total" enterprise rows are emitted. Two layouts:

  state_6digitnaics_xlsx  us_state_6digitnaics_2017.xlsx-style workbooks
                          (State, State Name, NAICS, NAICS Description,
                           Enterprise Size, Firms, Establishments, Employment, ...)
  us_receiptsize_xlsx     us_6digitnaics_r_2012.xlsx-style US-only workbooks
                          (NAICS CODE, NAICS DESCRIPTION, ENTERPRISE RECEIPT SIZE,
                           NUMBER OF FIRMS, NUMBER OF ESTABLISHMENTS, EMPLOYMENT, ...);
                          the 01: Total receipt-size band covers all employer
                          enterprises, so its firm counts equal the employment-size
                          file's 01: Total counts.

Used by StructuralVintageCalibration.jl (stage2b workstream 2b-4). Deliberately
stdlib+openpyxl only.
"""
import argparse
import csv
import sys

import openpyxl


def is_total_band(value):
    return value is not None and str(value).strip().startswith("01")


def convert_state_6digitnaics(worksheet, writer):
    emitted = 0
    for row in worksheet.iter_rows(values_only=True):
        if row[0] is None or str(row[0]).strip() != "00":
            continue
        if not is_total_band(row[4]):
            continue
        naics = str(row[2]).strip()
        writer.writerow(["00", naics, "01", row[5], row[6], row[7]])
        emitted += 1
    return emitted


def convert_us_receiptsize(worksheet, writer):
    emitted = 0
    for row in worksheet.iter_rows(values_only=True):
        if row[0] is None or row[2] is None:
            continue
        if not is_total_band(row[2]):
            continue
        naics = str(row[0]).strip()
        if naics.lower().startswith("naics"):
            continue
        writer.writerow(["00", naics, "01", row[3], row[4], row[5]])
        emitted += 1
    return emitted


def main():
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("--input", required=True)
    parser.add_argument(
        "--layout", required=True,
        choices=["state_6digitnaics_xlsx", "us_receiptsize_xlsx"],
    )
    parser.add_argument("--output", required=True)
    args = parser.parse_args()

    workbook = openpyxl.load_workbook(args.input, read_only=True)
    worksheet = workbook[workbook.sheetnames[0]]
    with open(args.output, "w", newline="") as handle:
        writer = csv.writer(handle)
        writer.writerow(["STATE", "NAICS", "ENTRSIZE", "FIRM", "ESTB", "EMPL"])
        if args.layout == "state_6digitnaics_xlsx":
            emitted = convert_state_6digitnaics(worksheet, writer)
        else:
            emitted = convert_us_receiptsize(worksheet, writer)
    if emitted < 500:
        print(f"ERROR: only {emitted} national total rows extracted", file=sys.stderr)
        sys.exit(1)
    print(f"wrote {emitted} national total rows to {args.output}")


if __name__ == "__main__":
    main()
